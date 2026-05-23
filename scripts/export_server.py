"""
Export sidecar: POST /export → XLSX from template
Run: uvicorn scripts.export_server:app --port 8001
  or: cd scripts && uvicorn export_server:app --port 8001
  or: python scripts/export_server.py
"""
from fastapi import FastAPI, HTTPException
from fastapi.responses import StreamingResponse
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
from typing import Any
import openpyxl
import io
import os
from datetime import datetime

app = FastAPI()
app.add_middleware(CORSMiddleware, allow_origins=["*"], allow_methods=["*"], allow_headers=["*"])

TEMPLATE_PATH = os.path.join(os.path.dirname(__file__), "..", "templates", "campaign-template.xlsx")

class ExportRequest(BaseModel):
    campaign: dict[str, Any]
    domains: list[dict[str, Any]]

# Columns we're allowed to write in the CM sheet
WRITABLE_COLUMNS = {
    "Period", "Period Start Date", "Order #", "Order Date", "Placement URL",
    "Order Price", "Target URL", "Anchor Text", "Link Type", "Budget",
    "Status", "Contact Email", "Team", "Notes", "Review Status",
    "Topics/Snippets", "GP Doc", "Content Status", "Payment Invoice", "Payment Status"
}

PROTECTED_SHEETS = {"__CM_HISTORY", "__CM_STATE"}

@app.post("/export")
async def export_xlsx(req: ExportRequest):
    if not os.path.exists(TEMPLATE_PATH):
        raise HTTPException(404, f"Template not found at {TEMPLATE_PATH}")

    wb = openpyxl.load_workbook(TEMPLATE_PATH)

    # Find CM sheet (try common names)
    cm_sheet = None
    for name in wb.sheetnames:
        if name in PROTECTED_SHEETS:
            continue
        if "CM" in name.upper() and "HISTORY" not in name.upper() and "STATE" not in name.upper():
            cm_sheet = wb[name]
            break
    if not cm_sheet:
        # Use first non-protected sheet
        for name in wb.sheetnames:
            if name not in PROTECTED_SHEETS:
                cm_sheet = wb[name]
                break

    if not cm_sheet:
        raise HTTPException(500, "No writable sheet found in template")

    # Build column index from header row (row 1)
    col_map = {}
    for col_idx in range(1, cm_sheet.max_column + 1):
        header = cm_sheet.cell(row=1, column=col_idx).value
        if header and str(header).strip() in WRITABLE_COLUMNS:
            col_map[str(header).strip()] = col_idx

    # Track formula cells in template (row 2+ patterns)
    formula_cells = set()
    for row in range(2, min(cm_sheet.max_row + 1, 10)):  # Check first few template rows
        for col in range(1, cm_sheet.max_column + 1):
            cell = cm_sheet.cell(row=row, column=col)
            if cell.data_type == 'f' or (isinstance(cell.value, str) and cell.value.startswith("=")):
                formula_cells.add(col)

    # Write domain data starting at row 2
    for i, domain in enumerate(req.domains):
        row_num = i + 2

        def safe_write(col_name: str, value: Any):
            if col_name not in col_map:
                return
            col_idx = col_map[col_name]
            # Don't overwrite formula columns
            if col_idx in formula_cells:
                existing = cm_sheet.cell(row=row_num, column=col_idx).value
                if isinstance(existing, str) and existing.startswith("="):
                    return
            cm_sheet.cell(row=row_num, column=col_idx).value = value

        safe_write("Placement URL", domain.get("domain", ""))
        safe_write("Order Price", domain.get("gp_price") or domain.get("li_price"))
        safe_write("Link Type", domain.get("link_type", ""))
        safe_write("Contact Email", domain.get("contact_email", ""))
        safe_write("Status", "Pending")
        safe_write("Order Date", datetime.now().strftime("%Y-%m-%d"))
        safe_write("Order #", i + 1)

    # Write to Client Info tab if it exists
    for name in wb.sheetnames:
        if "client" in name.lower() and "info" in name.lower():
            ci = wb[name]
            # Find first empty row
            write_row = ci.max_row + 1
            campaign = req.campaign
            ci.cell(row=write_row, column=1).value = campaign.get("clientName", "")
            ci.cell(row=write_row, column=2).value = campaign.get("niches", "")
            ci.cell(row=write_row, column=3).value = campaign.get("geo", "")
            ci.cell(row=write_row, column=4).value = campaign.get("profile", "")
            ci.cell(row=write_row, column=5).value = datetime.now().strftime("%Y-%m-%d")
            break

    # Write to Referring Domains tab if it exists
    for name in wb.sheetnames:
        if "referring" in name.lower() and "domain" in name.lower():
            rd = wb[name]
            for i, domain in enumerate(req.domains):
                row_num = i + 2  # after header
                rd.cell(row=row_num, column=1).value = i + 1
                rd.cell(row=row_num, column=2).value = domain.get("domain", "")
                rd.cell(row=row_num, column=3).value = domain.get("dr")
                rd.cell(row=row_num, column=4).value = domain.get("traffic")
                rd.cell(row=row_num, column=5).value = domain.get("totalScore", "")
            break

    # Save to buffer
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"campaign-export-{datetime.now().strftime('%Y%m%d-%H%M%S')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )

@app.get("/health")
async def health():
    return {"status": "ok", "template_exists": os.path.exists(TEMPLATE_PATH)}

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8001)
