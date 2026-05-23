"""
Fidelity check: verify formulas are preserved after export
Usage: python scripts/fidelity_check.py
"""
import openpyxl
import os, sys, json, tempfile

TEMPLATE_PATH = os.path.join(os.path.dirname(__file__), "..", "templates", "campaign-template.xlsx")

def check_fidelity():
    if not os.path.exists(TEMPLATE_PATH):
        print(f"FAIL: Template not found at {TEMPLATE_PATH}")
        return False

    # Load template, record all formula cells
    wb_template = openpyxl.load_workbook(TEMPLATE_PATH)
    formulas = {}

    for sheet_name in wb_template.sheetnames:
        ws = wb_template[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                if cell.data_type == 'f' or (isinstance(cell.value, str) and str(cell.value).startswith("=")):
                    key = f"{sheet_name}!{cell.coordinate}"
                    formulas[key] = cell.value

    print(f"Found {len(formulas)} formula cells in template")

    # Simulate an export by making the POST request
    import requests
    test_data = {
        "campaign": {"clientName": "Test Client", "niches": "test", "geo": "global", "profile": "standard"},
        "domains": [{"domain": "example.com", "dr": 50, "traffic": 5000, "gp_price": 100, "link_type": "GP", "contact_email": "test@test.com"}]
    }

    try:
        resp = requests.post("http://localhost:8001/export", json=test_data)
        if resp.status_code != 200:
            print(f"FAIL: Export returned {resp.status_code}")
            return False
    except Exception as e:
        print(f"FAIL: Could not reach export server: {e}")
        print("Make sure to run: python scripts/export_server.py")
        return False

    # Save response to temp file and reload
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp.write(resp.content)
        tmp_path = tmp.name

    wb_export = openpyxl.load_workbook(tmp_path)

    # Check every formula cell
    passed = 0
    failed = 0
    for key, original_formula in formulas.items():
        sheet_name, coord = key.split("!")
        if sheet_name not in wb_export.sheetnames:
            print(f"  WARN: Sheet {sheet_name} missing in export")
            continue
        cell = wb_export[sheet_name][coord]
        if cell.data_type == 'f' or (isinstance(cell.value, str) and str(cell.value).startswith("=")):
            passed += 1
        else:
            print(f"  FAIL: {key} was formula '{original_formula}' but is now '{cell.value}'")
            failed += 1

    os.unlink(tmp_path)

    print(f"\nFormula fidelity: {passed} passed, {failed} failed out of {len(formulas)}")
    if failed == 0:
        print("PASS: All formulas preserved")
        return True
    else:
        print("FAIL: Some formulas were overwritten")
        return False

if __name__ == "__main__":
    success = check_fidelity()
    sys.exit(0 if success else 1)
